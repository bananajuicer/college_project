from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QPushButton,
                             QLabel, QMessageBox, QLineEdit, QTableWidget,
                             QTableWidgetItem, QFileDialog, QComboBox, QInputDialog)
from db import DB
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import sqlite3

from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt


class AdminWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Панель администратора')  # Заголовок окна
        #self.setGeometry(100, 100, 400, 300)  # Размеры окна
        self.setMinimumSize(500, 800)
        self.setMaximumSize(600, 900)
        self.db = DB()


        # Основной Layout
        self.layout = QVBoxLayout()

        # Шрифт
        label_font = QFont("Arial", 14, QFont.Bold)
        button_font = QFont("Arial", 10)

        # Метка приветствия
        self.greeting_label = QLabel('Добро пожаловать, Администратор!')
        self.greeting_label.setFont(label_font)
        self.layout.addWidget(self.greeting_label)

        # Кнопка для добавления студента
        self.add_student_button = QPushButton('Добавить студента')
        self.add_student_button.setFont(button_font)
        self.add_student_button.clicked.connect(self.add_student)
        self.layout.addWidget(self.add_student_button)

        self.delete_student_button = QPushButton('Удалить студента')
        self.delete_student_button.setFont(button_font)
        self.delete_student_button.clicked.connect(self.delete_student)
        self.layout.addWidget(self.delete_student_button)

        # Кнопка для просмотра всех студентов
        self.view_students_button = QPushButton('Просмотреть всех студентов')
        self.view_students_button.setFont(button_font)
        self.view_students_button.clicked.connect(self.view_students)
        self.layout.addWidget(self.view_students_button)

        # Кнопка для добавления расписания
        self.add_student_schedule_button = QPushButton('Добавить расписание')
        self.add_student_schedule_button.setFont(button_font)
        self.add_student_schedule_button.clicked.connect(self.add_student_schedule)
        self.layout.addWidget(self.add_student_schedule_button)

        # Кнопки для работы с курсами
        self.add_course_button = QPushButton('Добавить курс')
        self.add_course_button.setFont(button_font)
        self.add_course_button.clicked.connect(self.add_course)
        self.layout.addWidget(self.add_course_button)

        self.delete_course_button = QPushButton('Удалить курс')
        self.delete_course_button.setFont(button_font)
        self.delete_course_button.clicked.connect(self.remove_course)
        self.layout.addWidget(self.delete_course_button)

        self.add_topic_button = QPushButton('Добавить тему')
        self.add_topic_button.setFont(button_font)
        self.add_topic_button.clicked.connect(self.add_topic)
        self.layout.addWidget(self.add_topic_button)

        self.view_courses_button = QPushButton('Получить все курсы')
        self.view_courses_button.setFont(button_font)
        self.view_courses_button.clicked.connect(self.view_courses)
        self.layout.addWidget(self.view_courses_button)

        self.add_study_materials_button = QPushButton('Добавить учебные материалы')
        self.add_study_materials_button.setFont(button_font)
        self.add_study_materials_button.clicked.connect(self.add_study_material)
        self.layout.addWidget(self.add_study_materials_button)

        # Кнопка для удаления всех материалов студента
        self.delete_student_materials_button = QPushButton('Удалить все материалы студента')
        self.delete_student_materials_button.setFont(button_font)
        self.delete_student_materials_button.clicked.connect(self.delete_student_materials)
        self.layout.addWidget(self.delete_student_materials_button)

        self.export_button = QPushButton('Экспортировать студентов в Excel')
        self.export_button.setFont(button_font)
        self.export_button.clicked.connect(self.export_to_excel)
        self.layout.addWidget(self.export_button)

        self.view_course_popularity_button = QPushButton('Анализ популярности курсов')
        self.view_course_popularity_button.setFont(button_font)
        self.view_course_popularity_button.clicked.connect(self.plot_course_popularity)
        self.layout.addWidget(self.view_course_popularity_button)

        # Кнопка выхода
        self.exit_button = QPushButton('Выход')
        self.exit_button.setFont(button_font)
        self.exit_button.clicked.connect(self.close)
        self.layout.addWidget(self.exit_button)

        # Устанавливаем layout
        self.setLayout(self.layout)
        self.apply_styles()

    def apply_styles(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
            }
            QLabel {
                color: #333333;
                margin: 10px 0;
            }
            QPushButton {
                background-color: #ffffff;
                border: 1px solid #cccccc;
                border-radius: 10px;
                padding: 10px;
                margin: 5px 0;
                color: #555555;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #e6f7ff;
            }
            QPushButton:pressed {
                background-color: #cceeff;
            }
        """)

    def add_student(self):
        # Создаем новое окно для ввода данных
        self.student_window = QWidget()
        self.student_window.setWindowTitle('Добавить студента')
        self.student_window.setGeometry(150, 150, 400, 300)

        layout = QVBoxLayout()
        layout.setSpacing(15)  # Отступы между элементами

        # Шрифты
        input_font = QFont("Arial", 12)
        button_font = QFont("Arial", 11, QFont.Bold)

        # Поля для ввода данных
        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText('Логин')
        self.login_input.setFont(input_font)
        layout.addWidget(self.login_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText('Пароль')
        self.password_input.setFont(input_font)
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        self.last_name_input = QLineEdit()
        self.last_name_input.setPlaceholderText('Фамилия')
        self.last_name_input.setFont(input_font)
        layout.addWidget(self.last_name_input)

        self.first_name_input = QLineEdit()
        self.first_name_input.setPlaceholderText('Имя')
        self.first_name_input.setFont(input_font)
        layout.addWidget(self.first_name_input)

        # Кнопка для добавления
        self.submit_button = QPushButton('Добавить')
        self.submit_button.setFont(button_font)
        self.submit_button.setCursor(Qt.PointingHandCursor)
        self.submit_button.clicked.connect(self.submit_student)
        layout.addWidget(self.submit_button)

        # Стилизация окна и элементов
        self.student_window.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;  /* Светлый фон */
            }
            QLineEdit {
                background-color: #ffffff;  /* Белый фон */
                border: 2px solid #d6d6d6;  /* Светлая рамка */
                border-radius: 8px;  /* Закругленные углы */
                padding: 8px 12px;  /* Внутренний отступ */
                color: #333333;  /* Цвет текста */
                font-size: 14px;  /* Размер шрифта */
            }
            QLineEdit:focus {
                border: 2px solid #0078d7;  /* Синяя рамка при фокусе */
                background-color: #eef6ff;  /* Светло-голубой фон */
            }
            QPushButton {
                background-color: #0078d7;  /* Синий фон кнопки */
                color: #ffffff;  /* Белый текст */
                border: none;
                border-radius: 8px;  /* Закругленные углы */
                padding: 10px 20px;  /* Внутренний отступ */
                font-size: 14px;  /* Размер шрифта */
                margin-top: 10px;  /* Отступ сверху */
            }
            QPushButton:hover {
                background-color: #005fa3;  /* Темно-синий фон при наведении */
            }
            QPushButton:pressed {
                background-color: #00457a;  /* Темно-синий фон при нажатии */
            }
        """)

        self.student_window.setLayout(layout)
        self.student_window.show()
    def submit_student(self):
        # Получаем данные из полей
        login = self.login_input.text().strip()
        password = self.password_input.text().strip()
        last_name = self.last_name_input.text().strip()
        first_name = self.first_name_input.text().strip()
        # Проверка на пустые поля
        if not login or not password or not last_name or not first_name:
            QMessageBox.warning(self, 'Ошибка', 'Все поля должны быть заполнены!')
            return
        # Проверка на формат логина
        import re
        username_pattern = r"^[a-zA-Z0-9_.-]+$"
        if not re.match(username_pattern, login):
            QMessageBox.warning(self, 'Ошибка',
                                'Логин может содержать только латинские буквы, цифры, точки, дефисы и символы подчёркивания!')
            return
        # Проверка на длину логина
        if len(login) < 3:
            QMessageBox.warning(self, 'Ошибка', 'Логин должен содержать не менее 3 символов!')
            return
        # Проверка на формат пароля
        password_pattern = r"^[a-zA-Z0-9@#$%^&+=]+$"  # Пароль может содержать только латинские буквы, цифры и символы @#$%^&+=
        if not re.match(password_pattern, password):
            QMessageBox.warning(self, 'Ошибка',
                                'Пароль может содержать только латинские буквы, цифры и символы @#$%^&+=!')
            return
        # Проверка длины пароля
        if len(password) < 6:
            QMessageBox.warning(self, 'Ошибка', 'Пароль должен содержать не менее 6 символов!')
            return
        # Проверка на формат имени и фамилии
        name_pattern = r"^[А-Яа-яЁёA-Za-z-]+$"  # Разрешены только буквы (латиница или кириллица) и дефисы
        if not re.match(name_pattern, last_name):
            QMessageBox.warning(self, 'Ошибка', 'Фамилия может содержать только буквы и дефисы!')
            return
        if not re.match(name_pattern, first_name):
            QMessageBox.warning(self, 'Ошибка', 'Имя может содержать только буквы и дефисы!')
            return
        # Добавление пользователя в базу данных
        try:
            self.db.add_user(login, password, last_name, first_name, 'student')  # статус 'student'
            QMessageBox.information(self, 'Успех', 'Студент успешно добавлен!')
            self.student_window.close()  # Закрываем окно добавления
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, 'Ошибка', 'Логин уже существует в системе!')
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    def add_student_schedule(self):
        # Создаем новое окно для ввода данных расписания
        self.schedule_window = QWidget()
        self.schedule_window.setWindowTitle('Добавить расписание')
        self.schedule_window.setGeometry(150, 150, 400, 400)

        layout = QVBoxLayout()
        layout.setSpacing(15)  # Пространство между элементами

        # Шрифты
        input_font = QFont("Arial", 12)
        button_font = QFont("Arial", 11, QFont.Bold)

        # Поля для ввода данных расписания
        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText('Формат даты: yyyy-MM-dd')
        self.date_input.setFont(input_font)
        layout.addWidget(self.date_input)

        self.day_input = QLineEdit()
        self.day_input.setPlaceholderText('День недели')
        self.day_input.setFont(input_font)
        layout.addWidget(self.day_input)

        self.subject_input = QLineEdit()
        self.subject_input.setPlaceholderText('Предмет')
        self.subject_input.setFont(input_font)
        layout.addWidget(self.subject_input)

        self.start_time_input = QLineEdit()
        self.start_time_input.setPlaceholderText('Время начала (чч:мм)')
        self.start_time_input.setFont(input_font)
        layout.addWidget(self.start_time_input)

        self.end_time_input = QLineEdit()
        self.end_time_input.setPlaceholderText('Время окончания (чч:мм)')
        self.end_time_input.setFont(input_font)
        layout.addWidget(self.end_time_input)

        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText('Логин студента')
        self.login_input.setFont(input_font)
        layout.addWidget(self.login_input)

        # Кнопка для добавления расписания
        self.submit_button = QPushButton('Добавить')
        self.submit_button.setFont(button_font)
        self.submit_button.setCursor(Qt.PointingHandCursor)
        self.submit_button.clicked.connect(self.submit_schedule)
        layout.addWidget(self.submit_button)

        # Стилизация окна и элементов
        self.schedule_window.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;  /* Светлый фон */
            }
            QLineEdit {
                background-color: #ffffff;  /* Белый фон */
                border: 2px solid #d6d6d6;  /* Серый контур */
                border-radius: 8px;  /* Закругленные углы */
                padding: 8px 12px;  /* Внутренние отступы */
                font-size: 14px;  /* Размер шрифта */
                color: #333333;  /* Цвет текста */
            }
            QLineEdit:focus {
                border: 2px solid #0078d7;  /* Синий контур при фокусе */
                background-color: #eef6ff;  /* Светло-голубой фон при фокусе */
            }
            QPushButton {
                background-color: #0078d7;  /* Синий фон кнопки */
                color: #ffffff;  /* Белый текст */
                border: none;
                border-radius: 8px;  /* Закругленные углы */
                padding: 10px 20px;  /* Внутренние отступы */
                font-size: 14px;  /* Размер шрифта */
                margin-top: 10px;  /* Отступ сверху */
            }
            QPushButton:hover {
                background-color: #005fa3;  /* Темно-синий фон при наведении */
            }
            QPushButton:pressed {
                background-color: #00457a;  /* Темно-синий фон при нажатии */
            }
        """)

        self.schedule_window.setLayout(layout)
        self.schedule_window.show()

    def submit_schedule(self):
        # Получаем данные из полей
        date = self.date_input.text().strip()
        day = self.day_input.text().strip()
        subject = self.subject_input.text().strip()
        time_start = self.start_time_input.text().strip()
        time_end = self.end_time_input.text().strip()
        login = self.login_input.text().strip()

        # Проверки на ввод данных
        if not date:
            QMessageBox.warning(self, 'Ошибка', 'Введите дату!')
            return
        if not day:
            QMessageBox.warning(self, 'Ошибка', 'Введите день недели!')
            return
        if not subject:
            QMessageBox.warning(self, 'Ошибка', 'Введите предмет!')
            return
        if not time_start:
            QMessageBox.warning(self, 'Ошибка', 'Введите время начала!')
            return
        if not time_end:
            QMessageBox.warning(self, 'Ошибка', 'Введите время окончания!')
            return
        if not login:
            QMessageBox.warning(self, 'Ошибка', 'Введите логин!')
            return

        # Дополнительные проверки
        # Проверка формата даты (например, YYYY-MM-DD)
        if not self.validate_date(date):
            QMessageBox.warning(self, 'Ошибка', 'Введите дату в формате ГГГГ-ММ-ДД!')
            return

        # Проверка формата времени (например, HH:MM)
        if not self.validate_time(time_start) or not self.validate_time(time_end):
            QMessageBox.warning(self, 'Ошибка', 'Введите время в формате ЧЧ:ММ!')
            return

        # Проверка существования пользователя
        user = self.db.get_user(login)
        if not user:
            QMessageBox.warning(self, 'Ошибка', 'Пользователь с таким логином не найден!')
            return

        # Проверка правильности времени (время начала должно быть раньше времени окончания)
        if time_start >= time_end:
            QMessageBox.warning(self, 'Ошибка', 'Время начала должно быть раньше времени окончания!')
            return

        # Добавляем расписание в базу данных
        try:
            self.db.add_student_schedule(date, day, subject, time_start, time_end, login)
            QMessageBox.information(self, 'Успех', 'Расписание успешно добавлено!')
            self.schedule_window.close()  # Закрываем окно добавления
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    def validate_date(self, date):
        """Проверяет, что дата соответствует формату ГГГГ-ММ-ДД."""
        import re
        return re.match(r"^\d{4}-\d{2}-\d{2}$", date) is not None

    def validate_time(self, time):
        """Проверяет, что время соответствует формату ЧЧ:ММ."""
        import re
        return re.match(r"^\d{2}:\d{2}$", time) is not None

    def view_students(self):
        # Создаем новое окно для просмотра студентов
        self.students_window = QWidget()
        self.students_window.setWindowTitle('Список студентов')
        self.students_window.setGeometry(150, 150, 500, 350)

        layout = QVBoxLayout()
        layout.setSpacing(15)  # Пространство между элементами

        # Создаем таблицу для отображения студентов
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(3)
        self.students_table.setHorizontalHeaderLabels(['Логин', 'Фамилия', 'Имя'])

        # Устанавливаем стиль таблицы
        self.students_table.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;  /* Белый фон таблицы */
                border-radius: 10px;  /* Закругленные углы */
                border: 1px solid #cccccc;  /* Легкая рамка */
            }
            QTableWidget::item {
                padding: 10px;  /* Отступы внутри ячеек */
                font-size: 14px;  /* Размер шрифта */
                color: #333333;  /* Цвет текста */
            }
            QHeaderView::section {
                background-color: #0078d7;  /* Синий фон для заголовков */
                color: #ffffff;  /* Белый цвет текста */
                font-size: 14px;  /* Размер шрифта заголовков */
                padding: 10px;  /* Отступы */
                text-align: center;  /* Выравнивание текста по центру */
            }
            QTableWidget::horizontalHeader {
                font-weight: bold;  /* Жирный шрифт для заголовков */
            }
            QTableWidget::item-selected {
                background-color: #eef6ff;  /* Легкий голубой фон для выбранных ячеек */
            }
        """)

        # Получаем список студентов из базы данных
        try:
            students = self.db.get_all_students()  # Предполагаем, что этот метод возвращает список студентов
            if not students:
                QMessageBox.information(self.students_window, 'Информация', 'Нет студентов для отображения.')
            else:
                self.students_table.setRowCount(len(students))  # Устанавливаем количество строк
                for row, student in enumerate(students):
                    # student - это кортеж (login, password, last_name, first_name)
                    self.students_table.setItem(row, 0, QTableWidgetItem(student[0]))  # Логин
                    self.students_table.setItem(row, 1, QTableWidgetItem(student[2]))  # Фамилия
                    self.students_table.setItem(row, 2, QTableWidgetItem(student[3]))  # Имя

                    # Делаем ячейки недоступными для редактирования
                    for col in range(3):
                        item = self.students_table.item(row, col)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Убираем флаг редактирования

        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

        layout.addWidget(self.students_table)
        self.students_window.setLayout(layout)
        self.students_window.show()

    def delete_student(self):
        # Создаем новое окно для ввода данных
        self.delete_window = QWidget()
        self.delete_window.setWindowTitle('Удалить студента')
        self.delete_window.setGeometry(150, 150, 300, 150)

        layout = QVBoxLayout()
        layout.setSpacing(20)  # Пространство между элементами

        # Поле для ввода логина студента для удаления
        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText('Введите логин студента')
        self.login_input.setFont(QFont("Inter", 12))  # Шрифт
        self.login_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.login_input)

        # Кнопка для удаления
        self.delete_button = QPushButton('Удалить')
        self.delete_button.setFont(QFont("Inter", 12))
        self.delete_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d7;
                color: white;
                padding: 10px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
            }
            QPushButton:hover {
                background-color: #005fa3;
            }
            QPushButton:pressed {
                background-color: #003f7f;
            }
        """)
        self.delete_button.clicked.connect(self.submit_delete)
        layout.addWidget(self.delete_button)

        # Устанавливаем layout
        self.delete_window.setLayout(layout)
        self.delete_window.show()

    def submit_delete(self):  # Удаляем студента из базы данных
        login = self.login_input.text().strip()

        if not login:  # Проверка на пустой ввод
            QMessageBox.warning(self, 'Ошибка', 'Введите логин для удаления!')
            return

        # Пытаемся удалить пользователя из базы данных
        try:
            user = self.db.get_user(login)  # Проверяем наличие пользователя
            if not user:
                QMessageBox.warning(self, 'Ошибка', 'Пользователь с таким логином не найден!')
                return

            self.db.remove_user(login)  # Удаляем пользователя
            QMessageBox.information(self, 'Успех', 'Студент успешно удалён!')
            self.delete_window.close()  # Закрываем окно удаления
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    def add_course(self):
        # Окно для добавления курса
        self.course_window = QWidget()
        self.course_window.setWindowTitle('Добавить курс')
        self.course_window.setGeometry(150, 150, 300, 200)

        layout = QVBoxLayout()
        layout.setSpacing(20)  # Пространство между элементами

        # Поле для ввода названия курса
        self.course_name_input = QLineEdit()
        self.course_name_input.setPlaceholderText('Название курса')
        self.course_name_input.setFont(QFont("Inter", 12))
        self.course_name_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.course_name_input)

        # Поле для ввода описания курса
        self.course_description_input = QLineEdit()
        self.course_description_input.setPlaceholderText('Описание курса')
        self.course_description_input.setFont(QFont("Inter", 12))
        self.course_description_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.course_description_input)

        # Кнопка для добавления курса
        self.submit_course_button = QPushButton('Добавить курс')
        self.submit_course_button.setFont(QFont("Inter", 12))
        self.submit_course_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d7;
                color: white;
                padding: 10px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
            }
            QPushButton:hover {
                background-color: #005fa3;
            }
            QPushButton:pressed {
                background-color: #003f7f;
            }
        """)
        self.submit_course_button.clicked.connect(self.submit_course)
        layout.addWidget(self.submit_course_button)

        # Устанавливаем layout
        self.course_window.setLayout(layout)
        self.course_window.show()

    def submit_course(self):
        # Получаем данные из полей
        name = self.course_name_input.text().strip()
        description = self.course_description_input.text().strip()

        # Проверки на ввод данных
        if not name:
            QMessageBox.warning(self, 'Ошибка', 'Введите название курса!')
            return
        if len(name) < 3:
            QMessageBox.warning(self, 'Ошибка', 'Название курса должно содержать минимум 3 символа!')
            return
        if not description:
            QMessageBox.warning(self, 'Ошибка', 'Введите описание курса!')
            return
        if len(description) < 10:
            QMessageBox.warning(self, 'Ошибка', 'Описание курса должно содержать минимум 10 символов!')
            return

        # Добавляем курс в базу данных
        try:
            self.db.add_course(name, description)
            QMessageBox.information(self, 'Успех', 'Курс успешно добавлен!')
            self.course_window.close()
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    def remove_course(self):
        # Окно для удаления курса
        self.remove_course_window = QWidget()
        self.remove_course_window.setWindowTitle('Удалить курс')
        self.remove_course_window.setGeometry(150, 150, 300, 200)

        layout = QVBoxLayout()
        layout.setSpacing(20)  # Пространство между элементами

        # Поле для ввода ID курса для удаления
        self.course_id_input = QLineEdit()
        self.course_id_input.setPlaceholderText('ID курса для удаления')
        self.course_id_input.setFont(QFont("Inter", 12))
        self.course_id_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.course_id_input)

        # Кнопка для удаления курса
        self.submit_remove_button = QPushButton('Удалить курс')
        self.submit_remove_button.setFont(QFont("Inter", 12))
        self.submit_remove_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                padding: 12px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #b03a2e;
            }
        """)
        self.submit_remove_button.clicked.connect(self.submit_remove_course)
        layout.addWidget(self.submit_remove_button)

        # Устанавливаем layout
        self.remove_course_window.setLayout(layout)
        self.remove_course_window.show()

    def submit_remove_course(self):
        # Получаем ID курса из поля ввода
        course_id = self.course_id_input.text().strip()

        # Проверяем ввод
        if not course_id:
            QMessageBox.warning(self.remove_course_window, 'Ошибка', 'Введите ID курса!')
            return
        if not course_id.isdigit():
            QMessageBox.warning(self.remove_course_window, 'Ошибка', 'ID курса должен быть числом!')
            return

        course_id = int(course_id)

        # Удаление курса из базы данных
        try:
            if not self.db.get_course_by_id(course_id):
                QMessageBox.warning(self.remove_course_window, 'Ошибка', 'Курс с таким ID не найден!')
                return

            self.db.delete_course(course_id)
            QMessageBox.information(self.remove_course_window, 'Успех', 'Курс успешно удалён!')
            self.remove_course_window.close()
        except Exception as e:
            QMessageBox.warning(self.remove_course_window, 'Ошибка', str(e))

    def add_topic(self):
        # Окно для добавления темы
        self.topic_window = QWidget()
        self.topic_window.setWindowTitle('Добавить тему')
        self.topic_window.setGeometry(150, 150, 300, 250)

        layout = QVBoxLayout()
        layout.setSpacing(20)  # Пространство между элементами

        # Поле для ввода ID курса
        self.topic_course_id_input = QLineEdit()
        self.topic_course_id_input.setPlaceholderText('ID курса')
        self.topic_course_id_input.setFont(QFont("Inter", 12))
        self.topic_course_id_input.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.topic_course_id_input)

        # Поле для ввода названия темы
        self.topic_name_input = QLineEdit()
        self.topic_name_input.setPlaceholderText('Название темы')
        self.topic_name_input.setFont(QFont("Inter", 12))
        self.topic_name_input.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.topic_name_input)

        # Поле для ввода содержания темы
        self.topic_content_input = QLineEdit()
        self.topic_content_input.setPlaceholderText('Содержание темы')
        self.topic_content_input.setFont(QFont("Inter", 12))
        self.topic_content_input.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border-radius: 10px;
                border: 1px solid #cccccc;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
                background-color: #ffffff;
            }
        """)
        layout.addWidget(self.topic_content_input)

        # Кнопка для добавления темы
        self.submit_topic_button = QPushButton('Добавить тему')
        self.submit_topic_button.setFont(QFont("Inter", 12))
        self.submit_topic_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;  /* Зеленый цвет */
                color: white;
                padding: 12px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
        """)
        self.submit_topic_button.clicked.connect(self.submit_topic)
        layout.addWidget(self.submit_topic_button)

        # Устанавливаем layout
        self.topic_window.setLayout(layout)
        self.topic_window.show()

    def submit_topic(self):
        # Получаем данные из полей
        course_id = self.topic_course_id_input.text().strip()
        name = self.topic_name_input.text().strip()
        content = self.topic_content_input.text().strip()

        # Проверки на ввод данных
        if not course_id:
            QMessageBox.warning(self, 'Ошибка', 'Введите ID курса!')
            return
        if not course_id.isdigit():
            QMessageBox.warning(self, 'Ошибка', 'ID курса должен быть числом!')
            return
        if not name:
            QMessageBox.warning(self, 'Ошибка', 'Введите название темы!')
            return
        if len(name) < 3:
            QMessageBox.warning(self, 'Ошибка', 'Название темы должно содержать минимум 3 символа!')
            return
        if not content:
            QMessageBox.warning(self, 'Ошибка', 'Введите содержание темы!')
            return
        if len(content) < 10:
            QMessageBox.warning(self, 'Ошибка', 'Содержание темы должно быть минимум из 10 символов!')
            return

        # Добавляем тему в базу данных
        try:
            self.db.add_topic(int(course_id), name, content)
            QMessageBox.information(self, 'Успех', 'Тема успешно добавлена!')
            self.topic_window.close()
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', str(e))

    def view_courses(self):
        # Окно для просмотра курсов
        self.courses_window = QWidget()
        self.courses_window.setWindowTitle('Список курсов')
        self.courses_window.setGeometry(150, 150, 500, 350)

        layout = QVBoxLayout()
        layout.setSpacing(20)  # Пространство между элементами

        # Создаем таблицу для отображения курсов
        self.courses_table = QTableWidget()
        self.courses_table.setColumnCount(3)
        self.courses_table.setHorizontalHeaderLabels(['ID', 'Название', 'Описание'])

        # Стиль таблицы
        self.courses_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                background-color: #fafafa;
                font-size: 14px;
                padding: 10px;
            }
            QHeaderView::section {
                background-color: #0078d7;
                color: white;
                font-size: 14px;
                padding: 8px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ddd;
            }
            QTableWidget::item:hover {
                background-color: #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #e0f7fa;
            }
        """)

        # Получаем список курсов из базы данных
        try:
            courses = self.db.get_all_courses()  # Предполагаем, что метод возвращает список курсов
            if not courses:
                QMessageBox.information(self.courses_window, 'Информация', 'Нет курсов для отображения.')
            else:
                self.courses_table.setRowCount(len(courses))  # Устанавливаем количество строк
                for row, course in enumerate(courses):
                    # course - это кортеж (id, name, description)
                    # Добавляем элементы в таблицу
                    self.courses_table.setItem(row, 0, QTableWidgetItem(str(course[0])))  # ID
                    self.courses_table.setItem(row, 1, QTableWidgetItem(course[1]))  # Название
                    self.courses_table.setItem(row, 2, QTableWidgetItem(course[2]))  # Описание

                    # Делаем ячейки недоступными для редактирования
                    for col in range(3):
                        item = self.courses_table.item(row, col)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Убираем флаг редактирования

        except Exception as e:
            QMessageBox.warning(self.courses_window, 'Ошибка', str(e))

        layout.addWidget(self.courses_table)

        # Добавление кнопки для закрытия окна
        close_button = QPushButton('Закрыть')
        close_button.setFont(QFont("Inter", 12))
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 12px;
                border-radius: 10px;
                font-size: 14px;
                border: none;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
        """)
        close_button.clicked.connect(self.courses_window.close)  # Закрытие окна при нажатии
        layout.addWidget(close_button)

        self.courses_window.setLayout(layout)
        self.courses_window.show()

    def add_study_material(self):
        # Получаем список всех курсов
        courses = self.db.get_all_courses()

        if not courses:
            QMessageBox.warning(self, 'Ошибка', 'Нет доступных курсов для добавления материалов.')
            return

        # Окно выбора курса
        course_dialog = QComboBox(self)
        for course in courses:
            course_dialog.addItem(course[1], userData=course[0])  # Добавляем название курса, а ID курса в userData

        # Окно выбора файлов
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("PDF и DOCX (*.pdf *.docx)")
        file_dialog.setViewMode(QFileDialog.List)

        if file_dialog.exec_():
            # Получаем путь к выбранным файлам
            files = file_dialog.selectedFiles()
            if files:
                file_path = files[0]  # Извлекаем путь к первому выбранному файлу
                file_name = file_path.split('/')[-1]  # Получаем имя файла
                file_type = file_path.split('.')[-1]  # Получаем расширение файла (pdf или docx)

                # Получаем выбранный курс
                selected_course_id = course_dialog.currentData()

                # Добавляем учебный материал в базу данных
                try:
                    self.db.add_study_material(selected_course_id, file_name, file_path, file_type)
                    QMessageBox.information(self, 'Успех', 'Учебный материал успешно добавлен!')
                except Exception as e:
                    QMessageBox.warning(self, 'Ошибка', f"Произошла ошибка при добавлении материала: {e}")
        else:
            QMessageBox.information(self, 'Информация', 'Файл не был выбран.')

        # Закрываем окно выбора курса, если файл не выбран или завершен процесс
        course_dialog.close()

    def export_to_excel(self):
        # Получаем данные для экспорта (пример: список студентов)
        try:
            students = self.db.get_all_students_1()
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось получить данные: {e}')
            return

        if not students:
            QMessageBox.information(self, 'Информация', 'Нет данных для экспорта.')
            return

        # Создаем новый Excel-файл
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Студенты"

        # Заголовки таблицы
        headers = ['Логин', 'Фамилия', 'Имя']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

        # Заполнение данных
        for row_num, student in enumerate(students, 2):  # Начиная со второй строки
            for col_num, value in enumerate(student, 1):  # Логин, фамилия, имя
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}{row_num}"] = value

        # Выбор файла для сохранения с названием по умолчанию
        options = QFileDialog.Options()
        default_name = "список всех студентов.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить как",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )

        if save_path:
            try:
                workbook.save(save_path)
                QMessageBox.information(self, 'Успех', f'Данные успешно сохранены в {save_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить файл: {e}')
    def plot_course_popularity(self):
        try:
            # Получаем данные о популярности курсов
            course_data = self.db.get_course_popularity()  # Предполагается, что метод возвращает список кортежей (course_name, student_count)

            if not course_data:
                QMessageBox.information(self, 'Информация', 'Нет данных для построения диаграммы.')
                return

            # Разбиваем данные на названия курсов и количество студентов
            course_names = [course[0] for course in course_data]
            student_counts = [course[1] for course in course_data]

            # Построение круговой диаграммы
            plt.figure(figsize=(8, 6))
            plt.pie(student_counts, labels=course_names, autopct='%1.1f%%', startangle=140)
            plt.title('Популярность курсов')
            plt.axis('equal')  # Убедимся, что круг не искажен
            plt.show()

        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f"Произошла ошибка при построении диаграммы: {e}")

    def delete_student_materials(self):
        # Окно выбора студента (по логину)
        student_login, ok = QInputDialog.getText(self, 'Удаление материалов', 'Введите логин студента:')
        if not ok or not student_login:
            return  # Выход, если не введен логин

        # Получаем список курсов, на которые записан студент
        self.db.cursor.execute('''
            SELECT course_id FROM student_courses WHERE student_login = ?
        ''', (student_login,))
        courses = self.db.cursor.fetchall()

        if not courses:
            QMessageBox.warning(self, 'Ошибка', 'Студент не записан ни на один курс.')
            return

        # Для каждого курса удаляем связанные учебные материалы
        for course in courses:
            course_id = course[0]
            self.db.delete_study_materials_for_course(course_id)  # Метод для удаления материалов по ID курса

        QMessageBox.information(self, 'Успех', 'Все материалы студента успешно удалены.')